"""
Shared utilities for Streamlit UI pages.

Provides common functionality for:
- Results table formatting (currency, percentages, commas)
- Fabric Data Warehouse connection management
- Session state initialization
"""

from __future__ import annotations

import io
import re

import pandas as pd
import streamlit as st

from tools.fabric_conn import FabricConnection

# ---------------------------------------------------------------------------
# Results Formatting
# ---------------------------------------------------------------------------

# Column names (lowercased) that represent dollar amounts
_CURRENCY_COLS = {"cost", "profit", "revenue", "spend", "cpc", "exchange revenue"}
# Column names (lowercased) that represent rates / percentages
_RATE_COLS = {"ctr", "conversion rate", "click through rate", "roi"}

# Rate metrics that must be recalculated from base sums rather than summed directly.
# Maps rate keyword → (numerator keyword, denominator keyword) — all lowercased.
_RATE_FORMULAS = {
    "ctr": ("clicks", "impressions"),
    "click through rate": ("clicks", "impressions"),
    "conversion rate": ("conversions", "clicks"),
    "cpc": ("cost", "clicks"),
    "roi": ("profit", "cost"),
    "revenue per click": ("revenue", "clicks"),
    "revenue per conversion": ("revenue", "conversions"),
}


def format_results(df: pd.DataFrame):
    """
    Return a pandas Styler with currency, comma, and percentage formatting.

    Args:
        df: Raw DataFrame from query execution

    Returns:
        Styled DataFrame with formatted numeric columns, or raw df if no numeric columns
    """
    # Create a copy to avoid modifying the original
    df_display = df.copy()
    fmt: dict[str, str] = {}

    for col in df_display.columns:
        if not pd.api.types.is_numeric_dtype(df_display[col]):
            continue
        col_lower = col.lower().replace("_", " ").strip()
        if any(kw in col_lower for kw in _CURRENCY_COLS):
            fmt[col] = "${:,.2f}"
        elif any(kw in col_lower for kw in _RATE_COLS):
            # Multiply rate columns by 100 for percentage display
            df_display[col] = df_display[col] * 100
            fmt[col] = "{:,.2f}%"
        elif pd.api.types.is_integer_dtype(df_display[col]):
            fmt[col] = "{:,}"
        else:
            fmt[col] = "{:,.2f}"
    return df_display.style.format(fmt) if fmt else df_display


def build_totals_row(df: pd.DataFrame) -> pd.DataFrame | None:
    """
    Build a single 'Total' summary row for numeric columns.

    Base metrics (cost, clicks, etc.) are summed. Rate/percentage metrics
    (CTR, conversion rate, etc.) are recalculated from the summed bases so
    the aggregate is correct rather than an average of row-level rates.

    Returns None when there are ≤1 rows (totals would be redundant) or when
    no numeric columns exist.
    """
    if df is None or len(df) <= 1:
        return None

    numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    if not numeric_cols:
        return None

    col_sums = {col: df[col].sum() for col in numeric_cols}
    totals: dict = {}

    # Rate keyword detection — matches MDR's rate_keywords list
    _rate_kws = ["rate", "percentage", "percent", "share", "roi", "ctr", "cpc",
                 "revenue per click", "revenue per conversion"]

    # Fill categorical columns with a label
    # Show "Total" only in the first dimension column, blank in others for cleaner display
    first_dim_col = None
    for col in df.columns:
        if not pd.api.types.is_numeric_dtype(df[col]):
            if first_dim_col is None:
                first_dim_col = col
                totals[col] = "Total"
            else:
                totals[col] = ""  # Blank for other dimension columns

    # For each numeric column decide: sum, recalculate from base metrics, or mean
    recalculated = set()
    for col in numeric_cols:
        col_lower = col.lower().replace("_", " ").strip()
        for rate_kw, (num_kw, denom_kw) in _RATE_FORMULAS.items():
            if rate_kw in col_lower:
                # Recalculate from summed base metrics (most accurate)
                num_col = next(
                    (c for c in numeric_cols if num_kw in c.lower().replace("_", " ")), None
                )
                denom_col = next(
                    (c for c in numeric_cols if denom_kw in c.lower().replace("_", " ")), None
                )
                if num_col and denom_col and col_sums.get(denom_col, 0) != 0:
                    totals[col] = col_sums[num_col] / col_sums[denom_col]
                else:
                    # Base metrics not in result — use mean (matches MDR fallback behavior)
                    totals[col] = df[col].mean()
                recalculated.add(col)
                break

        if col not in recalculated:
            col_lower = col.lower().replace("_", " ").strip()
            if any(kw in col_lower for kw in _rate_kws):
                # Unknown rate metric — mean is safer than sum
                totals[col] = df[col].mean()
            else:
                totals[col] = col_sums[col]

    return pd.DataFrame([totals], columns=df.columns)


# ---------------------------------------------------------------------------
# Export utilities
# ---------------------------------------------------------------------------

def sanitize_filename(text: str, max_len: int = 50) -> str:
    """Convert arbitrary text into a safe filename stem (no extension).

    Strips non-word characters, collapses whitespace to underscores, and
    truncates to max_len.  Falls back to "export" for empty input.
    """
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE)
    text = re.sub(r"\s+", "_", text.strip())
    return text[:max_len].strip("_") or "export"


def build_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Build an in-memory Excel workbook with one sheet per DataFrame.

    Applies column number formatting:
      - Currency columns ($#,##0.00)
      - Rate/percentage columns (0.00%)
      - Integer columns (#,##0)
      - Other numeric columns (#,##0.00)

    Args:
        sheets: Mapping of sheet_name → DataFrame.  DataFrames with a named
                index (e.g., MDR summary matrix) are reset so the index
                becomes the first column.

    Returns:
        Raw bytes of the .xlsx file, ready for st.download_button().
    """
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            # Promote named index to a regular column (e.g., MDR "Date Range" index)
            df_out = df.reset_index() if df.index.name else df.copy()
            safe_sheet = sheet_name[:31]  # Excel sheet-name limit
            df_out.to_excel(writer, sheet_name=safe_sheet, index=False)

            ws = writer.sheets[safe_sheet]

            # Apply per-column number format + auto-width
            for col_idx, col_name in enumerate(df_out.columns, start=1):
                col_letter = get_column_letter(col_idx)
                col_lower = str(col_name).lower().replace("_", " ").strip()

                # Auto-width: longest value or header, capped at 30
                col_width = max(
                    len(str(col_name)),
                    max((len(str(ws.cell(r, col_idx).value or "")) for r in range(2, ws.max_row + 1)), default=0),
                )
                ws.column_dimensions[col_letter].width = min(col_width + 2, 30)

                if not pd.api.types.is_numeric_dtype(df_out[col_name]):
                    continue  # no number format for text columns

                if any(kw in col_lower for kw in _CURRENCY_COLS):
                    num_fmt = '$#,##0.00'
                elif any(kw in col_lower for kw in _RATE_COLS):
                    num_fmt = '0.00%'
                elif pd.api.types.is_integer_dtype(df_out[col_name]):
                    num_fmt = '#,##0'
                else:
                    num_fmt = '#,##0.00'

                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = num_fmt

    return buf.getvalue()


# ---------------------------------------------------------------------------
# Fabric Connection UI
# ---------------------------------------------------------------------------

def init_fabric_state():
    """Initialize Fabric connection session state keys if not present."""
    if "fabric_conn" not in st.session_state:
        st.session_state.fabric_conn = None
    if "fabric_connected" not in st.session_state:
        st.session_state.fabric_connected = False


def render_fabric_sidebar():
    """
    Render the Fabric Data Warehouse connection section in the sidebar.

    Shows current connection status and Connect/Disconnect buttons.
    Manages st.session_state.fabric_conn and st.session_state.fabric_connected.
    """
    st.subheader("Fabric Connection")
    fc: FabricConnection | None = st.session_state.fabric_conn

    # Show current connection status
    if st.session_state.fabric_connected and fc is not None:
        st.success("Fabric: connected")
        st.caption(f"Server: {fc.server[:30]}...")
        st.caption(f"Database: {fc.database}")
        if st.button("Disconnect", use_container_width=True):
            fc.close()
            st.session_state.fabric_connected = False
            st.rerun()
    else:
        st.warning("Fabric: not connected")
        if st.button("Connect to Fabric", type="primary", use_container_width=True):
            with st.spinner("Authenticating via Azure AD..."):
                try:
                    conn = FabricConnection()
                    conn.connect()
                    st.session_state.fabric_conn = conn
                    st.session_state.fabric_connected = True
                    st.rerun()
                except Exception as e:
                    st.error(f"Connection failed: {e}")
